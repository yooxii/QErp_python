import openpyxl as xl
import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import sys
import json
import DealTxt as dt

def load_config():
    """加载配置文件"""
    root = tk.Tk()
    root.withdraw()

    try:
        cfgPath = filedialog.askopenfilename(title='选择配置文件', filetypes=[('JSON', '*.json')])
        if not cfgPath:
            raise FileNotFoundError("未选择配置文件")
        with open(cfgPath, 'r', encoding='utf-8') as f:
            qerp = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError) as e:
        messagebox.showerror("错误", f"加载配置文件失败: {str(e)}")
        sys.exit()
    finally:
        root.destroy()
    
    return cfgPath, qerp

def create_tk_root(title, width, height, x_shift=0, y_shift=0):
    """创建一个Tkinter根窗口"""
    root = tk.Tk()
    root.title(title)
    root.geometry(f"{width}x{height}+{x_shift}+{y_shift}")
    return root

def cfg_excel(root_main, qerp):
    """配置Excel窗口"""
    root = create_tk_root("配置Excel", 500, 300, (root_main.winfo_screenwidth() - 500) // 2, (root_main.winfo_screenheight() - 300) // 2)

    input_box = create_report_inputbox(root, qerp['Report'])
    
    btn_frame = tk.Frame(root)
    btn_frame.pack(pady=20)
    
    btn_save = tk.Button(btn_frame, text="应用", command=lambda: (save_reportcfg(input_box, qerp), root.destroy(), initialize_window(root_main, qerp)))
    btn_save.pack(side=tk.LEFT, padx=10)

    btn_exit = tk.Button(btn_frame, text="返回", command=root.destroy)
    btn_exit.pack(side=tk.LEFT, padx=10)

    root.mainloop()

def cfg_txt(root_main,qerp):
    """配置TXT窗口"""
    root = create_tk_root("配置TXT", 500, 500, (root_main.winfo_screenwidth() - 500) // 2, (root_main.winfo_screenheight() - 500) // 2)

    txtcfg_canvas = tk.Canvas(root, width=480, height=480)
    txtcfg_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    scrollbar = tk.Scrollbar(root, orient=tk.VERTICAL, command=txtcfg_canvas.yview)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    txtcfg_canvas.configure(yscrollcommand=scrollbar.set)
    txtcfg_canvas.bind("<Configure>", lambda e: txtcfg_canvas.configure(scrollregion=txtcfg_canvas.bbox("all")))

    data_frame = tk.Frame(txtcfg_canvas)
    txtcfg_canvas.create_window((0, 0), window=data_frame, anchor="nw")

    input_frame = tk.Frame(data_frame)
    input_frame.grid(row=0, column=0, padx=10, pady=10)

    input_box = create_txt_inputbox(input_frame, qerp['TXT'])
    
    btn_frame = tk.Frame(data_frame)
    btn_frame.grid(row=1, column=0, padx=10, pady=10)

    btn_save = tk.Button(btn_frame, text="应用", command=lambda: (save_txtcfg(input_box, qerp), root.destroy()))
    btn_save.pack(side=tk.LEFT, padx=10)

    btn_exit = tk.Button(btn_frame, text="返回", command=root.destroy)
    btn_exit.pack(side=tk.LEFT, padx=10)

    root.mainloop()

def cfg_select(root, qerp):
    """配置Select窗口"""

    if qerp['Report'] == {}:
        messagebox.showerror("错误", "请先配置报告！")
        return

    root = create_tk_root("配置Select", 600, 600, (root_main.winfo_screenwidth() - 500) // 2, (root_main.winfo_screenheight() - 300) // 2)

    selectcfg_canvas = tk.Canvas(root, width=540, height=560)
    selectcfg_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    scrollbar = tk.Scrollbar(root, orient=tk.VERTICAL, command=selectcfg_canvas.yview)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    selectcfg_canvas.configure(yscrollcommand=scrollbar.set)
    selectcfg_canvas.bind("<Configure>", lambda e: selectcfg_canvas.configure(scrollregion=selectcfg_canvas.bbox("all")))

    data_frame = tk.Frame(selectcfg_canvas)
    selectcfg_canvas.create_window((0, 0), window=data_frame, anchor="nw")

    input_frame = tk.Frame(data_frame)
    input_frame.grid(row=0, column=0, padx=10, pady=10)

    input_box = create_select_inputbox(input_frame, qerp['Select'])
    
    btn_frame = tk.Frame(data_frame)
    btn_frame.grid(row=1, column=0, padx=10, pady=10)
    
    btn_save = tk.Button(btn_frame, text="应用", command=lambda: (save_selectcfg(input_box, qerp), root.destroy()))
    btn_save.pack(side=tk.LEFT, padx=10)

    btn_exit = tk.Button(btn_frame, text="返回", command=root.destroy)
    btn_exit.pack(side=tk.LEFT, padx=10)

    root.mainloop()

def save_reportcfg(input_box, qerp):
    """保存报告配置"""
    report = qerp['Report']
    for i, input_excel in enumerate(input_box):
        report[list(report.keys())[i]] = input_excel.get()
    savecfg(qerp)

def save_txtcfg(input_box, qerp):
    """保存数据文件配置"""
    txt = qerp['TXT']
    for i, input_txt in enumerate(input_box):
        # 如果input_txt是ENTRY，则直接获取值
        if isinstance(txt[list(txt.keys())[i]], str):
            txt[list(txt.keys())[i]] = input_txt.get()
        # 否则用json.loads解析
        else:
            txt[list(txt.keys())[i]] = json.loads(input_txt.get(1.0, tk.END))
    savecfg(qerp)

def save_selectcfg(input_box, qerp):
    """保存选择配置"""
    select = qerp['Select']
    for i, input_select in enumerate(input_box):
        select[list(select.keys())[i]] = json.loads(input_select.get(1.0, tk.END))
    savecfg(qerp)

def export_cfg(qerp):
    """导出配置"""
    # 选择导出路径
    exportPath = filedialog.asksaveasfilename(
        title="导出配置文件",
        initialdir=qerp['initialdir'],
        filetypes=[('JSON', '*.json')]
    )
    if not exportPath:
        return
    try:
        with open(f"{exportPath}.bak", 'w', encoding='utf-8') as f:
            json.dump(qerp, f, ensure_ascii=False, indent=4)
        messagebox.showinfo("提示", "导出成功！")
    except IOError as e:
        messagebox.showerror("错误", f"导出配置文件失败: {str(e)}")

def import_cfg(qerp, cfgPath):
    """导入配置"""
    # 选择导入路径
    importPath = filedialog.askopenfilename(
        title="导入配置文件",
        initialdir=qerp['initialdir'],
        filetypes=[('JSON', '*.json')]
    )
    if not importPath:
        return
    try:
        with open(importPath, 'r', encoding='utf-8') as f:
            qerp_import = json.load(f)
        qerp.update(qerp_import)
        messagebox.showinfo("提示", "导入成功！")
    except (FileNotFoundError, json.JSONDecodeError) as e:
        messagebox.showerror("错误", f"导入配置文件失败: {str(e)}")

def savecfg(qerp):
    try:
        with open(cfgPath, 'w', encoding='utf-8') as f:
            json.dump(qerp, f, ensure_ascii=False, indent=4)
    except IOError as e:
        messagebox.showerror("错误", f"保存配置文件失败: {str(e)}")

def show_datas(qerp):
    """显示数据窗口"""    
    root = tk.Tk()
    root.title("QErp")

    global root_main
    root_main = root

    WidthScreen = root.winfo_screenwidth()
    HeightScreen = root.winfo_screenheight()

    WidthMain = 600
    HeightMain = 600
    
    root.geometry(f"{WidthMain}x{HeightMain}+{int((WidthScreen-WidthMain)/2)}+{int((HeightScreen-HeightMain)/2)}")
    
    menubar = tk.Menu(root)

    file_menu = tk.Menu(menubar, tearoff=0)
    file_menu.add_command(label="选择文件夹", command=lambda: load_rootpath(qerp))
    file_menu.add_command(label="打开报告", command=lambda: initialize_window(root, qerp))
    file_menu.add_command(label="报告另存为", command=lambda: excelrp.save_excel(data_box, txt_seqs, qerp))
    menubar.add_cascade(label="报告", menu=file_menu)
    
    data_menu = tk.Menu(menubar, tearoff=0)
    data_menu.add_command(label="打开数据文件", command=lambda: load_tests(data_box))
    data_menu.add_command(label="加载选择", command=lambda: load_selects(data_box, qerp))
    data_menu.add_command(label="保存选择", command=lambda: save_select(data_box, qerp))
    menubar.add_cascade(label="数据处理", menu=data_menu)

    setting_menu = tk.Menu(menubar, tearoff=0)
    setting_menu.add_command(label="配置Report", command=lambda: cfg_excel(root, qerp))
    setting_menu.add_command(label="配置TXT", command=lambda: cfg_txt(root, qerp))
    setting_menu.add_command(label="配置Select", command=lambda: cfg_select(root, qerp))
    setting_menu.add_command(label="导出配置", command=lambda: export_cfg(qerp))
    setting_menu.add_command(label="导入配置", command=lambda: import_cfg(qerp, cfgPath))
    menubar.add_cascade(label="设置", menu=setting_menu)

    menubar.add_command(label="退出", command=root.quit)
    root.config(menu=menubar)

    root.mainloop()

def initialize_window(root, qerp):
    """初始化窗口布局和变量"""
    for widget in root.winfo_children():
        if not isinstance(widget, tk.Menu):
            widget.destroy()

    report = qerp['Report']

    global data_box, txt_seqs, excelrp, tests_name
    data_box = []
    txt_seqs = []
    excelrp = ExcelRp()
    tests_name = excelrp.read_excel(report)
    
    data_canvas = tk.Canvas(root, width=580, height=600)
    data_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    scrollbar = tk.Scrollbar(root, orient=tk.VERTICAL, command=data_canvas.yview)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    data_canvas.configure(yscrollcommand=scrollbar.set)
    data_canvas.bind("<Configure>", lambda e: data_canvas.configure(scrollregion=data_canvas.bbox("all")))

    data_frame = tk.Frame(data_canvas)
    data_canvas.create_window((0, 0), window=data_frame, anchor="nw")

    for key, value in tests_name.items():
        small_frame = tk.Frame(data_frame)
        small_frame.pack(pady=2, side=tk.TOP)
        tk.Label(small_frame, text=f"{key}:", justify='left', width=30).pack(anchor='w', pady=2, side=tk.LEFT)
        Comb_data = ttk.Combobox(small_frame, width=40, justify='center')
        Comb_data.insert(0, value)
        Comb_data.pack(pady=2, side=tk.LEFT)
        data_box.append(Comb_data)

def create_report_inputbox(root, report):
    """创建输入框"""
    input_box = []
    input_frame = tk.Frame(root)
    input_frame.pack(pady=2)
    
    for key, value in report.items():
        tk.Label(input_frame, text=f"{key}:").pack(anchor='w', pady=2)
        input_excel = tk.Entry(input_frame, width=40, justify='center')
        input_excel.insert(0, value)
        input_excel.pack(pady=2)
        input_box.append(input_excel)
    
    return input_box

def create_txt_inputbox(root, txt):
    """创建输入框"""
    input_box = []
    input_frame = tk.Frame(root)
    input_frame.pack(pady=2)
    
    for i, (key, value) in enumerate(txt.items()):
        tk.Label(input_frame, text=f"{key}:").grid(row=i, column=0, padx=10, pady=10)
        # 如果value是字符串直接输出
        if isinstance(value, str):
            input_excel = tk.Entry(input_frame, width=40, justify='center')
            input_excel.insert(0, value)
        # 否则用json.dumps输出
        else:
            height = len(value) + 2 if isinstance(value, list) else 2
            input_excel = tk.Text(input_frame, width=40, height=height, wrap='word')
            input_excel.insert(tk.INSERT, json.dumps(value, ensure_ascii=False, indent=4))
        input_excel.grid(row=i, column=1, padx=10, pady=10)
        input_box.append(input_excel)
    
    return input_box

def create_select_inputbox(root, selects):
    """创建输入框"""
    input_box = []
    input_frame = tk.Frame(root)
    input_frame.pack(pady=2)
    
    for i, (key, value) in enumerate(selects.items()):
        tk.Label(input_frame, text=f"{key}:").grid(row=i, column=0, padx=10, pady=10)
        height = len(value) + 2 if isinstance(value, list) else 2
        input_select = tk.Text(input_frame, width=50, height=height, wrap='word')
        input_select.insert(tk.INSERT, json.dumps(value, ensure_ascii=False, indent=4))
        input_select.grid(row=i, column=1, padx=10, pady=10)

        input_box.append(input_select)
    
    return input_box

def load_rootpath(qerp):
    """加载文件夹"""
    # 打开路径选择框
    rootpath = filedialog.askdirectory(
        title="选择文件夹",
        initialdir=qerp['initialdir']
    )
    if rootpath:
        qerp['initialdir'] = rootpath
    return rootpath

def load_tests(select_box):
    """加载数据"""
    global txt_seqs
    try:
        txt_seqs = dt.deal_data2(dt.deal_data1(dt.open_file()))
        if not txt_seqs or not txt_seqs[0]:
            raise ValueError("未找到有效的数据值")

        values = ["NA"] + list(txt_seqs[0].keys())
        for select in select_box:
            select['values'] = values
            select.set(values[0])

    except Exception as e:
        messagebox.showerror("错误", f"加载选择时出错: {str(e)}")

def load_selects(select_box, qerp):
    selects = qerp['Select']
    # 选取txt['select']的每一个元素的第一个，如果是字符串，则设置select_box的对应位置的值
    for i, select_key in enumerate(selects):
        if isinstance(selects[select_key], str):
            select_box[i].set(selects[select_key])
        else:
            select_box[i].set(selects[select_key][0])

def save_select(data_box, qerp):
    """保存选择"""
    selects = qerp['Select']
    if not data_box:
        messagebox.showwarning("警告", "请先加载数据！")
        return

    saveSelects = {key: data.get() for key, data in zip(tests_name.keys(), data_box)}

    for select_key, select_val in saveSelects.items():
        try:
            if isinstance(selects[select_key], list):
                selects[select_key][0] = select_val
            else:
                selects[select_key] = [select_val]
        except KeyError:
            selects[select_key] = [select_val]
    savecfg(qerp)

def find_tests_name(sheet, report):
    """找到测试项目名称和起始位置"""
    res = {}
    start = None
    
    for col in sheet.columns:
        for cell in col:
            if cell.value == report['flag_data_start_row']:
                r = cell.row
                for c in range(cell.column, sheet.max_column + 1):
                    if sheet.cell(row=r, column=c).value == report['flag_data_start_col']:
                        start = {'row': r, 'col': c}
                        break
                if start: break

    if start:
        for row in range(start['row'] + 1, sheet.max_row + 1):
            cellValue = sheet.cell(row=row, column=start['col']).value
            if cellValue is not None and sheet.cell(row=row, column=start['col'] + 1).value is None:
                res[cellValue] = {'row': row, 'col': start['col']}

    return res

def disp_save_select(data_box, qerp):
    """保存选择窗口"""
    root = create_tk_root("保存选择", 250, 150, (root_main.winfo_screenwidth() - 250) // 2, (root_main.winfo_screenheight() - 150) // 2)

    save_frame = tk.Frame(root)
    save_frame.pack(pady=10)
    
    tk.Label(save_frame, text="是否需要覆盖之前的选择？").pack(pady=20)

    tk.Button(save_frame, text="是", width=5, command=lambda: (save_select(data_box, qerp), root.destroy())).pack(side=tk.LEFT, padx=30)
    tk.Button(save_frame, text="否", width=5, command=root.destroy).pack(side=tk.LEFT, padx=30)

    root.mainloop()

class ExcelRp:
    """ExcelRp类，用于操作Excel文件"""

    def __init__(self):
        self.file_path = None
        self.wb = None

    def open_file(self, qerp):
        try:
            self.file_path = filedialog.askopenfilename(
                title="Open Excel file",
                initialdir=qerp['initialdir'],
                filetypes=[("Excel files", "*.xlsx *.xls")]
            )
            if self.file_path:
                self.wb = xl.load_workbook(self.file_path)
        except Exception as e:
            messagebox.showerror("错误", f"打开Excel文件失败: {str(e)}")
            self.wb = None

    def read_excel(self, report):
        if not self.file_path or not self.wb:
            self.open_file(qerp)
        if not self.wb:
            return {}
        sheet = self.wb[report['sheet_name']]
        res = find_tests_name(sheet, report)
        return res

    def save_excel(self, data_box, seqs, qerp):
        if not seqs:
            messagebox.showwarning("警告", "请先加载数据！")
            return

        save_file = filedialog.asksaveasfilename(
            title="另存为Excel文件",
            initialdir=qerp['initialdir'],
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if not save_file:
            return
        
        selects = qerp['Select']
        report = qerp['Report']

        saveSelects = {key: data.get() for key, data in zip(tests_name.keys(), data_box)}

        for i in range(len(seqs)):
            for select_key, select in saveSelects.items():
                if select == "NA":
                    continue
                if select_key in tests_name:
                    row = tests_name[select_key]['row']
                    col = tests_name[select_key]['col'] + i + 1
                    value = seqs[i].get(select, "")
                    flag = False
                    for st_key, st_value in selects.items():
                        if select_key.find(st_key) != -1:
                            Cvalue = value[st_value[1]][st_value[2]]
                            flag = True
                            break
                    if not flag:
                        # 未找到匹配的选择项，使用第一个值
                        Cvalue = value[list(value.keys())[0]][0]
                    # 保留三位小数
                    Cell = self.wb[report['sheet_name']].cell(row=row, column=col)
                    Cell.number_format = '0.000'
                    Cell.value = float(Cvalue)

        try:
            self.wb.save(f"{save_file.replace('.xlsx', '')}.xlsx")
            disp_save_select(data_box, qerp)
        except Exception as e:
            messagebox.showerror("错误", f"保存Excel文件失败: {str(e)}")

if __name__ == '__main__':
    cfgPath, qerp = load_config()
    show_datas(qerp)
